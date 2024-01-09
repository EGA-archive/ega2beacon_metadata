#!/usr/bin/env python

"""ega2beacon.py  :  Converts EGA metadata into Beacon metadata """

__author__ = "Raul Garcia, Mauricio Moldes"
__version__ = "0.1"
__maintainer__ = "Raul Garcia, Mauricio Moldes"
__email__ = "raul.garcia@crg.es,mauricio.moldes@crg.es"
__status__ = "development"

import logging
import yaml
import psycopg2
import sys
#from xml.dom.minidom import parse, parseString
from xml.dom import minidom
from openpyxl import load_workbook

logger = logging.getLogger('ega2beacon_logger')

#global my_cool_variable

""" VERIFIES THE CONNECTION TO PLSQL """
def connection_plsql(cfg):
    conn_string = "host='" + str(cfg['plsql']['host']) + "' dbname='" + str(
        cfg['plsql']['dbname']) + "' user='" + str(
        cfg['plsql']['user']) + "' password='" + str(cfg['plsql']['password']) + "' port = '" + str(
        cfg['plsql']['port']) + "'"
    conn_plsql = psycopg2.connect(conn_string)
    return conn_plsql


"""" Get metadata from PLSQL """
def get_metadata_egapro(conn_plsql):
    cursor = conn_plsql.cursor()
    sql = """SELECT ebi_xml
        FROM analysis_table at
        WHERE ega_stable_id = 'EGAZ00001450756'"""
    cursor.execute(sql)
    records = cursor.fetchone()
    return records

""" Parse xml EGA metadata """
def parse_xml_analysis(xml):
    dom = minidom.parseString(xml)  # parse data
    study_element = dom.getElementsByTagName('STUDY_REF')
 #   print(study_element)

    stable_id = str(study_element[0].attributes['accession'].value)
#    print(stable_id)

    try:
        sample_element = dom.getElementsByTagName('SAMPLE_REF')
     #   print(sample_element)
        sample_ref = str(sample_element[0].attributes['accession'].value)
      #  print(sample_ref)
    except:
        print('NA')

    return stable_id, sample_ref
 #   parsed_doc = xml.dom.minidom.parseString(xml_tree)

    #get root element
  #  root = parsed_doc.documentElement

    #get element by tag or id

    # my_cool_variable = stable_id



""" Convert EGA metadata into Beacon Friendly Format """
def convert_ega_beacon_metadata():


    return False



""" Write Beacon friendly format metadata """
def write_metadata_beacon(sample_ref):

    models = load_workbook('/Users/raul/Library/CloudStorage/OneDrive-CRG-CentredeRegulacioGenomica/ega.nosync/bioteam/ega2beacon/ega2beacon_metadata/files/Beacon-v2-Models_template.xlsx')

    sheet = models['analyses']

    # find  column index in the header
    header = sheet[1]
    try:
        biosample_id_column = header.index('biosampleId') + 1
    except Exception as e:
        print(error)
        return False

        next_row = sheet.max_row + 1

        # write sample_ref below header
        sheet.cell(row=next_row, column=biosample_id_column, value=sample_ref)

        # Save the modified Excel file
        workbook.save('/Users/raul/Library/CloudStorage/OneDrive-CRG-CentredeRegulacioGenomica/ega.nosync/bioteam/ega2beacon/Beacon-v2-Models_template_modified.xlsx')
        print("saved successfully")
    except Exception as e:
        print(error)
        return False

    return True






def ega2beacon():
    conn_plsql = None
    try:
        conn_plsql = connection_plsql(cfg)
        if conn_plsql:
            results = get_metadata_egapro(conn_plsql)
            xml = results[0]
            print(xml)

            stable_id, sample_ref = parse_xml_analysis(xml)
            print(stable_id, sample_ref)
           # print(my_cool_variable)
    except Exception as e:
        print(e)

""" READ CONFIG FILE """
def read_config():
    with open("../config/config.yml", 'r') as ymlfile:
        cfg = yaml.safe_load(ymlfile)
    return cfg

""" MAIN"""
def main():
    try:
        # configure logging
        log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s [in %(pathname)s:%(lineno)d]'
        logging.basicConfig(format=log_format)
        global cfg
        cfg = read_config()
        ega2beacon()

    except Exception as e:
        logger.error("Error: {}".format(e))
        sys.exit(-1)


if __name__ == '__main__':
    main()